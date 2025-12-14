"""
PHASE 7 - Tests TDD pour l'export Excel.
"""

import pytest
import pandas as pd
from pathlib import Path
import openpyxl


class TestWriteWorkbook:
    """Tests pour write_workbook()."""
    
    def test_creates_file(self, temp_dir):
        """write_workbook crée un fichier Excel."""
        from prism_salesops.export_xlsx import write_workbook
        
        sheets = {
            "Test": pd.DataFrame({"A": [1, 2, 3], "B": [4, 5, 6]})
        }
        
        path = temp_dir / "output.xlsx"
        write_workbook(path, sheets)
        
        assert path.exists()
    
    def test_creates_all_sheets(self, temp_dir):
        """write_workbook crée toutes les feuilles demandées."""
        from prism_salesops.export_xlsx import write_workbook
        
        sheets = {
            "Données enrichies": pd.DataFrame({"A": [1]}),
            "Détail": pd.DataFrame({"B": [2]}),
            "Top10": pd.DataFrame({"C": [3]}),
            "Résumé rep": pd.DataFrame({"D": [4]}),
            "Contrôles": pd.DataFrame({"E": [5]}),
            "Non appariés": pd.DataFrame({"F": [6]}),
        }
        
        path = temp_dir / "output.xlsx"
        write_workbook(path, sheets)
        
        # Vérifier les feuilles
        wb = openpyxl.load_workbook(path)
        sheet_names = wb.sheetnames
        
        for name in sheets.keys():
            assert name in sheet_names, f"Feuille manquante: {name}"
    
    def test_preserves_row_counts(self, temp_dir):
        """write_workbook préserve le nombre de lignes."""
        from prism_salesops.export_xlsx import write_workbook
        
        df = pd.DataFrame({"A": range(100), "B": range(100, 200)})
        sheets = {"Data": df}
        
        path = temp_dir / "output.xlsx"
        write_workbook(path, sheets)
        
        # Relire et vérifier
        read_df = pd.read_excel(path, sheet_name="Data")
        assert len(read_df) == 100
    
    def test_handles_empty_dataframe(self, temp_dir):
        """write_workbook gère les DataFrames vides."""
        from prism_salesops.export_xlsx import write_workbook
        
        sheets = {
            "Vide": pd.DataFrame(),
            "Non vide": pd.DataFrame({"A": [1, 2]}),
        }
        
        path = temp_dir / "output.xlsx"
        write_workbook(path, sheets)
        
        # Ne doit pas lever d'erreur
        wb = openpyxl.load_workbook(path)
        assert "Vide" in wb.sheetnames
    
    def test_handles_special_characters(self, temp_dir):
        """write_workbook gère les caractères spéciaux dans les données."""
        from prism_salesops.export_xlsx import write_workbook
        
        df = pd.DataFrame({
            "Nom": ["Café", "Résumé", "Naïf"],
            "Symboles": ["<test>", "A & B", '"Quotes"'],
        })
        sheets = {"Special": df}
        
        path = temp_dir / "output.xlsx"
        write_workbook(path, sheets)
        
        read_df = pd.read_excel(path, sheet_name="Special")
        assert "Café" in read_df["Nom"].values
    
    def test_overwrites_existing_file(self, temp_dir):
        """write_workbook écrase un fichier existant."""
        from prism_salesops.export_xlsx import write_workbook
        
        path = temp_dir / "output.xlsx"
        
        # Premier fichier
        sheets1 = {"Sheet1": pd.DataFrame({"A": [1]})}
        write_workbook(path, sheets1)
        
        # Deuxième fichier (écrase)
        sheets2 = {"Sheet2": pd.DataFrame({"B": [2]})}
        write_workbook(path, sheets2)
        
        wb = openpyxl.load_workbook(path)
        assert "Sheet2" in wb.sheetnames
        # Sheet1 ne doit plus exister (fichier écrasé)
        assert "Sheet1" not in wb.sheetnames
